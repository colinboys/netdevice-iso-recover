#!/usr/bin/env python3
"""parse_script.py — 解析隔离/恢复操作脚本，提取接口的 shutdown / undo shutdown 操作。

支持的厂家（按厂家分发）：
  - huawei  ：VRP V800（已实现：interface / shutdown / undo shutdown、// 段注释）
  - h3c     ：Comware V7（已实现：interface / shutdown / undo shutdown、Excel 第一列脚本）
  - zte     ：占位（待实现；中兴常用 no shutdown 而非 undo shutdown）
  - ruijie  ：占位（待实现；锐捷常用 no shutdown）

输出：JSON 数组，按脚本中出现顺序，每个元素：
  {
    "line_no":     原脚本 1-based 行号
    "iface":       操作的接口名
    "op":          "shutdown" | "undo_shutdown" | "invalid_shutdown_typo" | "invalid_undo_shutdown_typo"
    "section":     顶层注释（"//..."），无则 null
    "raw":         原始行
  }

使用：
  python3 parse_script.py <script> --vendor huawei
  python3 parse_script.py <script.xlsx> --vendor h3c --sheet 隔离脚本

注意：「no shutdown」在最终输出里统一归一为 "undo_shutdown"。拼写错误
（如 `shutdwon`）会被记录为 invalid_*，用于报告严重命令错误，不作为有效
端口操作覆盖。
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional


# ---------- 通用正则 ----------

IFACE_RE = re.compile(r"^\s*interface\s+(.+?)\s*$", re.IGNORECASE)
H3C_IFACE_RE = re.compile(r"^\s*interface\s*(.+?)\s*$", re.IGNORECASE)

# 区分三种「与关闭端口相关」的命令，避免短路误判
_RE_SHUTDOWN = re.compile(r"^\s*shutdown\s*$", re.IGNORECASE)
_RE_SHUTDOWN_TYPO = re.compile(r"^\s*shutdwon\s*$", re.IGNORECASE)
_RE_UNDO_SHUTDOWN = re.compile(r"^\s*undo\s+shutdown\s*$", re.IGNORECASE)
_RE_UNDO_SHUTDOWN_TYPO = re.compile(r"^\s*undo\s+shutdwon\s*$", re.IGNORECASE)
_RE_NO_SHUTDOWN = re.compile(r"^\s*no\s+shutdown\s*$", re.IGNORECASE)

# 顶层段注释（//...）— VRP 风格
SECTION_COMMENT_RE = re.compile(r"^\s*//\s*(?P<text>.+?)\s*$")


def _normalize_iface(name: str) -> str:
    """归一化接口名用于比较：去多余空白、保留单个空格、不改变大小写。

    注意：保留原始大小写以兼容『Eth-trunk 2』与『Eth-Trunk2』两种写法，
    比较时在调用方做大小写不敏感处理。
    """
    return " ".join(name.split())


def _edit_distance(a: str, b: str) -> int:
    """Return a small Levenshtein distance for command typo detection."""
    if a == b:
        return 0
    previous = list(range(len(b) + 1))
    for i, ca in enumerate(a, start=1):
        current = [i]
        for j, cb in enumerate(b, start=1):
            current.append(min(
                previous[j] + 1,
                current[j - 1] + 1,
                previous[j - 1] + (ca != cb),
            ))
        previous = current
    return previous[-1]


def _command_word(value: str) -> str:
    return re.sub(r"[^a-z]", "", value.lower())


def _looks_like_shutdown_typo(value: str) -> bool:
    word = _command_word(value)
    if not word or word == "shutdown":
        return False
    return word.startswith("shut") or _edit_distance(word, "shutdown") <= 2


def _looks_like_undo_typo(value: str) -> bool:
    word = _command_word(value)
    if not word or word in {"undo", "no"}:
        return False
    return word.startswith("un") or _edit_distance(word, "undo") <= 2


def _classify_shutdown_line(line: str) -> Optional[str]:
    """把当前行映射为 shutdown/undo_shutdown；命中不到返回 None。"""
    if _RE_UNDO_SHUTDOWN.match(line):
        return "undo_shutdown"
    if _RE_UNDO_SHUTDOWN_TYPO.match(line):
        return "invalid_undo_shutdown_typo"
    if _RE_NO_SHUTDOWN.match(line):
        # Cisco / 中兴 / 锐捷风格的 no shutdown，归一为 undo_shutdown
        return "undo_shutdown"
    if _RE_SHUTDOWN.match(line):
        return "shutdown"
    if _RE_SHUTDOWN_TYPO.match(line):
        return "invalid_shutdown_typo"
    words = line.split()
    if len(words) == 1 and _looks_like_shutdown_typo(words[0]):
        return "invalid_shutdown_typo"
    if len(words) == 2:
        first, second = words
        if (_RE_SHUTDOWN.match(second) or _looks_like_shutdown_typo(second)) and _looks_like_undo_typo(first):
            return "invalid_undo_shutdown_typo"
        if _command_word(first + second) == "shutdown":
            return "invalid_shutdown_typo"
    return None


def read_script_text(
    path: str | Path,
    preferred_sheet: Optional[str] = None,
    device_name: Optional[str] = None,
) -> str:
    """读取文本脚本或 Excel 脚本，返回按行拼接的命令文本。

    H3C 厂家样例常以 xlsx 交付：第一列是命令/段落，第二列是端口描述。
    解析时只取第一列；如果提供 preferred_sheet，优先读取同名工作表。
    """
    p = Path(path)
    if p.suffix.lower() in (".xlsx", ".xlsm"):
        return _read_xlsx_first_column(p, preferred_sheet, device_name)
    return p.read_text(encoding="utf-8", errors="replace")


def _read_xlsx_first_column(
    path: Path,
    preferred_sheet: Optional[str] = None,
    device_name: Optional[str] = None,
) -> str:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "读取 Excel 脚本需要 openpyxl。请安装 openpyxl，或将脚本另存为 txt 后再运行。"
        ) from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    if preferred_sheet and preferred_sheet in wb.sheetnames:
        ws = wb[preferred_sheet]
    else:
        ws = wb[wb.sheetnames[0]]

    raw_lines: List[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        value = row[0]
        if value is None:
            raw_lines.append("")
            continue
        raw_lines.append(str(value).strip())

    if not device_name:
        return "\n".join(raw_lines)

    start: Optional[int] = None
    for idx, line in enumerate(raw_lines):
        if device_name in line:
            start = idx
            break
    if start is None:
        return "\n".join(raw_lines)

    block: List[str] = []
    for line in raw_lines[start:]:
        if block and "NFV-D-" in line and device_name not in line and not line.lower().startswith("interface"):
            break
        block.append(line)
    return "\n".join(block)


# ---------- 华为 VRP（已实现） ----------

def parse_huawei(text: str) -> List[Dict[str, Any]]:
    """华为 VRP 风格的隔离/恢复脚本解析。

    识别要点：
      - `interface <ifname>` 进入接口上下文
      - `shutdown` / `undo shutdown` 记录为 op
      - 顶层 `//` 注释作为「section」分段提示
      - `#`、`!` 开头的行视为注释
      - `system-view` / `commit` / `save` / `quit` 等控制命令忽略
    """
    ops: List[Dict[str, Any]] = []
    current_iface: Optional[str] = None
    current_section: Optional[str] = None

    for idx, raw in enumerate(text.splitlines(), start=1):
        line = raw.rstrip("\n")

        # 顶层注释：//开头，且行首没有更深的缩进
        m = SECTION_COMMENT_RE.match(line)
        if m and not line.startswith(("  //", "\t//")):
            current_section = m.group("text").strip()
            current_iface = None  # 注释也会重置 current_iface，与 VRP 行为一致
            continue

        # 注释行：# 或 ! 开头
        if line.lstrip().startswith(("#", "!")):
            continue

        m_iface = IFACE_RE.match(line)
        if m_iface:
            current_iface = _normalize_iface(m_iface.group(1))
            continue

        if current_iface is None:
            # 等待进入 interface 段
            continue

        op = _classify_shutdown_line(line)
        if op is None:
            # 其它命令（commit / save / quit / system-view 等）忽略
            continue

        ops.append({
            "line_no": idx,
            "iface": current_iface,
            "op": op,
            "section": current_section,
            "raw": line.strip(),
        })

    return ops


# ---------- 中兴（占位） ----------

def parse_zte(text: str) -> List[Dict[str, Any]]:
    """中兴 ZXR10 / ROSNG 隔离/恢复脚本解析（占位）。"""
    raise NotImplementedError(
        "中兴（zte）隔离/恢复脚本解析尚未实现。\n"
        "实现前请先在 references/parser-zte.md 补充『configure terminal / interface / shutdown / no shutdown』等样例，"
        "再在本文件实现 parse_zte()。\n"
        "提示：中兴常用『shutdown』/『no shutdown』而非『undo shutdown』，"
        "no shutdown 已在 _classify_shutdown_line 中归一为 undo_shutdown。"
    )


# ---------- 华三 Comware V7 ----------

def parse_h3c(text: str) -> List[Dict[str, Any]]:
    """华三 Comware V7 风格脚本解析。

    识别要点：
      - `interface Route-Aggregation1` 与现场常见的
        `interfaceRoute-Aggregation1` 均可识别
      - `shutdown` / `undo shutdown` / `no shutdown`
      - 厂家 Excel 样例中的常见错拼 `shutdwon` 识别为非法命令，
        便于报告指出严重拼写错误
      - 中文段落行（如“关闭上行端口”）作为 section 保留
    """
    ops: List[Dict[str, Any]] = []
    current_iface: Optional[str] = None
    current_section: Optional[str] = None

    for idx, raw in enumerate(text.splitlines(), start=1):
        line = raw.rstrip("\n")
        stripped = line.strip()
        if not stripped:
            current_iface = None
            continue
        if stripped.startswith(("#", "!")):
            continue

        m_iface = H3C_IFACE_RE.match(stripped)
        if m_iface:
            current_iface = _normalize_iface(m_iface.group(1))
            continue

        op = _classify_shutdown_line(stripped)
        if current_iface is not None and op is not None:
            ops.append({
                "line_no": idx,
                "iface": current_iface,
                "op": op,
                "section": current_section,
                "raw": stripped,
            })
            continue

        # 控制命令忽略；其它非命令行视为段落提示。
        if stripped.lower() in {"quit", "return", "save", "system-view"}:
            current_iface = None if stripped.lower() == "quit" else current_iface
            continue
        if not re.search(r"\s", stripped) or any(ch >= "\u4e00" and ch <= "\u9fff" for ch in stripped):
            current_section = stripped

    return ops


# ---------- 锐捷（占位） ----------

def parse_ruijie(text: str) -> List[Dict[str, Any]]:
    """锐捷 RGOS 隔离/恢复脚本解析（占位）。"""
    raise NotImplementedError(
        "锐捷（ruijie）隔离/恢复脚本解析尚未实现。\n"
        "实现前请先在 references/parser-ruijie.md 补充『configure terminal / interface / shutdown / no shutdown』等样例，"
        "再在本文件实现 parse_ruijie()。\n"
        "提示：锐捷常用『shutdown』/『no shutdown』而非『undo shutdown』；"
        "接口名常含空格（如『GigabitEthernet 0/1』），IFACE_RE 的非贪婪匹配已能容忍此情形。"
    )


# ---------- 调度器 ----------

PARSERS: Dict[str, Callable[[str], List[Dict[str, Any]]]] = {
    "huawei": parse_huawei,
    "zte": parse_zte,
    "h3c": parse_h3c,
    "ruijie": parse_ruijie,
}


def parse_script(text: str, vendor: str = "huawei") -> List[Dict[str, Any]]:
    """按厂家分发的脚本解析入口。

    为保持向后兼容，未传 vendor 时默认按 huawei 处理（与改造前一致）。
    """
    key = (vendor or "huawei").lower()
    if key not in PARSERS:
        raise ValueError(
            f"未知厂商 {vendor!r}；当前支持：{', '.join(PARSERS.keys())}"
        )
    return PARSERS[key](text)


# ---------- CLI ----------

def main(argv: Iterable[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="解析隔离/恢复脚本中的接口操作")
    parser.add_argument("path", help="脚本文件路径")
    parser.add_argument(
        "--vendor",
        default="huawei",
        choices=sorted(PARSERS.keys()),
        help="设备厂家（默认 huawei）；中兴/华三/锐捷目前为占位，会抛 NotImplementedError",
    )
    parser.add_argument("--indent", type=int, default=2, help="JSON 缩进")
    parser.add_argument("--sheet", default=None, help="Excel 脚本工作表名（如 隔离脚本 / 恢复脚本）")
    args = parser.parse_args(list(argv) if argv is not None else None)

    text = read_script_text(args.path, preferred_sheet=args.sheet)
    ops = parse_script(text, args.vendor)
    json.dump(ops, sys.stdout, ensure_ascii=False, indent=args.indent)
    sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
